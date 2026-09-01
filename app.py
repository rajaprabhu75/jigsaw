import os
import time
import sqlite3
from io import BytesIO
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
DATABASE = os.path.join(BASE_DIR, "leaderboard.db")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

EVENT_NAME = "Piping Harmony 2026"

app = Flask(__name__)
app.secret_key = "change-this-secret-key-before-deploying"  # change in production
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024  # 8 MB max upload

# CHANGE THIS before sharing the app with anyone else.
# Admin panel lives at /admin/login — it is not linked from the player page.
ADMIN_PASSWORD = "admin123"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS leaderboard (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_number TEXT,
            player_name TEXT NOT NULL,
            image_name TEXT,
            grid_size TEXT,
            time_taken REAL,
            completed_at TEXT
        )
        """
    )
    # Every player who enters their name and starts is logged here immediately,
    # whether or not they go on to finish the puzzle.
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_number TEXT,
            player_name TEXT NOT NULL,
            image_name TEXT,
            grid_size TEXT,
            started_at TEXT
        )
        """
    )
    # Single-row table holding the puzzle the admin has currently set up.
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS puzzle_config (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            image_name TEXT,
            grid_size INTEGER,
            updated_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_active_config():
    conn = get_db()
    row = conn.execute("SELECT * FROM puzzle_config WHERE id = 1").fetchone()
    conn.close()
    return row


def set_active_config(image_name, grid_size):
    conn = get_db()
    conn.execute(
        """
        INSERT INTO puzzle_config (id, image_name, grid_size, updated_at)
        VALUES (1, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            image_name = excluded.image_name,
            grid_size = excluded.grid_size,
            updated_at = excluded.updated_at
        """,
        (image_name, grid_size, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Player-facing routes
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    config = get_active_config()

    if request.method == "POST":
        player_name = request.form.get("player_name", "").strip()
        reg_number = request.form.get("reg_number", "").strip()

        if not config or not config["image_name"]:
            flash("The admin hasn't set up a puzzle yet. Please check back soon.")
            return redirect(url_for("index"))

        if not player_name or not reg_number:
            flash("Please enter both your name and registration number.")
            return redirect(url_for("index"))

        conn = get_db()
        already_played = conn.execute(
            "SELECT 1 FROM registrations WHERE UPPER(reg_number) = UPPER(?) LIMIT 1",
            (reg_number,),
        ).fetchone()

        if already_played:
            conn.close()
            flash("This registration number has already played the puzzle. Each participant can only play once.")
            return redirect(url_for("index"))

        # Snapshot the currently active puzzle for this player's session,
        # so it stays consistent even if the admin changes it mid-game.
        session["player_name"] = player_name
        session["image_name"] = config["image_name"]
        session["grid_size"] = config["grid_size"]
        session["reg_number"] = reg_number

        # Log the registration immediately so the admin panel shows the
        # player's name as soon as they start, not only once they finish.
        conn.execute(
            "INSERT INTO registrations (reg_number, player_name, image_name, grid_size, started_at) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                reg_number,
                player_name,
                config["image_name"],
                f"{config['grid_size']}x{config['grid_size']}",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
        conn.close()

        return redirect(url_for("puzzle"))

    return render_template("index.html", config=config, event_name=EVENT_NAME)


@app.route("/puzzle")
def puzzle():
    if "player_name" not in session or "image_name" not in session:
        return redirect(url_for("index"))

    return render_template(
        "puzzle.html",
        player_name=session["player_name"],
        image_name=session["image_name"],
        grid_size=int(session.get("grid_size", 3)),
        reg_number=session.get("reg_number", "—"),
        event_name=EVENT_NAME,
    )


@app.route("/api/complete", methods=["POST"])
def complete():
    if "player_name" not in session:
        return jsonify({"error": "No active session"}), 400

    data = request.get_json(force=True, silent=True) or {}
    time_taken = data.get("time_taken")

    if time_taken is None:
        return jsonify({"error": "Missing time_taken"}), 400

    grid_size = session.get("grid_size", 3)

    conn = get_db()
    conn.execute(
        "INSERT INTO leaderboard (reg_number, player_name, image_name, grid_size, time_taken, completed_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            session.get("reg_number"),
            session["player_name"],
            session.get("image_name"),
            f"{grid_size}x{grid_size}",
            time_taken,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    conn.commit()
    conn.close()

    return jsonify({"status": "ok"})


@app.route("/play-again")
def play_again():
    session.pop("player_name", None)
    session.pop("image_name", None)
    session.pop("grid_size", None)
    session.pop("reg_number", None)
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# Admin routes (not linked anywhere on the player pages — go to /admin/login directly)
# ---------------------------------------------------------------------------

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin"))
        flash("Incorrect password.")
    return render_template("admin_login.html", event_name=EVENT_NAME)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/export")
def admin_export():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    conn = get_db()
    registrations = conn.execute("SELECT * FROM registrations ORDER BY id ASC").fetchall()
    leaderboard = conn.execute("SELECT * FROM leaderboard ORDER BY time_taken ASC").fetchall()
    conn.close()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Registrations"
    ws1.append(["Reg No", "Player Name", "Grid", "Started At"])
    for row in registrations:
        ws1.append([row["reg_number"], row["player_name"], row["grid_size"], row["started_at"]])

    ws2 = wb.create_sheet("Leaderboard")
    ws2.append(["Reg No", "Player Name", "Grid", "Time (s)", "Completed At", "Image"])
    for row in leaderboard:
        ws2.append([
            row["reg_number"],
            row["player_name"],
            row["grid_size"],
            row["time_taken"],
            row["completed_at"],
            row["image_name"],
        ])

    for ws in (ws1, ws2):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"piping_harmony_2026_puzzle_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/admin/reset-players")
def admin_reset_players():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    conn = get_db()
    conn.execute("DELETE FROM registrations")
    conn.execute("DELETE FROM leaderboard")
    conn.commit()
    conn.close()

    flash("All player registrations and leaderboard entries have been cleared.")
    return redirect(url_for("admin"))


@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        grid_size = request.form.get("grid_size", "3")
        image = request.files.get("image")
        config = get_active_config()

        image_name = config["image_name"] if config else None

        if image and image.filename != "":
            if not allowed_file(image.filename):
                flash("Please upload a valid image file (png, jpg, jpeg, gif, webp).")
                return redirect(url_for("admin"))
            filename = secure_filename(image.filename)
            image_name = f"{int(time.time())}_{filename}"
            image.save(os.path.join(app.config["UPLOAD_FOLDER"], image_name))
        elif not image_name:
            flash("Please upload an image for the puzzle.")
            return redirect(url_for("admin"))

        set_active_config(image_name, int(grid_size))
        flash("Puzzle updated — new players will now get this image and difficulty.")
        return redirect(url_for("admin"))

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM leaderboard ORDER BY time_taken ASC"
    ).fetchall()
    registrations = conn.execute(
        "SELECT * FROM registrations ORDER BY id DESC"
    ).fetchall()
    total_players = conn.execute(
        "SELECT COUNT(DISTINCT player_name) as c FROM leaderboard"
    ).fetchone()["c"]
    total_games = conn.execute(
        "SELECT COUNT(*) as c FROM leaderboard"
    ).fetchone()["c"]
    total_registrations = conn.execute(
        "SELECT COUNT(*) as c FROM registrations"
    ).fetchone()["c"]
    conn.close()

    config = get_active_config()

    return render_template(
        "admin.html",
        rows=rows,
        registrations=registrations,
        total_players=total_players,
        total_games=total_games,
        total_registrations=total_registrations,
        config=config,
        event_name=EVENT_NAME,
    )


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
